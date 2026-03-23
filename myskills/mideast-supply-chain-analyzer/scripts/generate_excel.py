#!/usr/bin/env python3
"""
generate_excel.py — 中东供应链风险 Excel 报表生成器

用途:
  接收完整的风险分析数据（JSON），生成格式化的 .xlsx 报表。
  包含4个 Sheet：新闻汇总、风险评分矩阵、供应链影响评估、应对建议。

用法:
  python generate_excel.py --input scored_data.json --output report.xlsx

依赖:
  pip install openpyxl
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装依赖：pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ─── 样式常量 ────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(
    bottom=Side(style="thin", color="FFFFFF"),
    right=Side(style="thin", color="FFFFFF"),
)

BODY_FONT = Font(name="Microsoft YaHei", size=10)
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)
BODY_BORDER = Border(
    bottom=Side(style="thin", color="E5E6EB"),
    right=Side(style="thin", color="E5E6EB"),
)

RISK_FILLS = {
    "高危": PatternFill(start_color="FF4D4F", end_color="FF4D4F", fill_type="solid"),
    "中等": PatternFill(start_color="FAAD14", end_color="FAAD14", fill_type="solid"),
    "低":   PatternFill(start_color="52C41A", end_color="52C41A", fill_type="solid"),
}

RISK_FONTS = {
    "高危": Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF"),
    "中等": Font(name="Microsoft YaHei", size=10, bold=True, color="1D2129"),
    "低":   Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF"),
}

LINK_FONT = Font(name="Microsoft YaHei", size=10, color="1677FF", underline="single")

SUMMARY_FILL = PatternFill(start_color="F0F5FF", end_color="F0F5FF", fill_type="solid")
SUMMARY_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color="1F4E79")


# ─── 工具函数 ────────────────────────────────────────────────────────────────

def apply_header_row(ws, headers: list, row: int = 1):
    """为指定行设置表头样式"""
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = HEADER_BORDER


def apply_body_cell(ws, row: int, col: int, value, align=None):
    """设置正文单元格样式"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BODY_FONT
    cell.alignment = align or BODY_ALIGN
    cell.border = BODY_BORDER
    return cell


def auto_column_width(ws, min_width: int = 10, max_width: int = 50):
    """根据内容自动调整列宽"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                # 中文字符占约2个字符宽度
                val_str = str(cell.value)
                char_len = 0
                for ch in val_str:
                    if ord(ch) > 127:
                        char_len += 2
                    else:
                        char_len += 1
                max_len = max(max_len, char_len)
        adjusted = max(min(max_len + 2, max_width), min_width)
        ws.column_dimensions[col_letter].width = adjusted


def apply_risk_fill(cell, risk_level: str):
    """根据风险等级设置单元格背景色"""
    if risk_level in RISK_FILLS:
        cell.fill = RISK_FILLS[risk_level]
        cell.font = RISK_FONTS[risk_level]


# ─── Sheet 1: 新闻汇总 ──────────────────────────────────────────────────────

def create_news_sheet(wb: Workbook, news_items: list):
    """Sheet1: 新闻汇总 — 时间/标题/来源/摘要/链接/语言"""
    ws = wb.active
    ws.title = "新闻汇总"

    headers = ["序号", "发布时间", "标题", "来源", "摘要", "链接", "语言"]
    apply_header_row(ws, headers)

    for i, item in enumerate(news_items, 1):
        row = i + 1
        apply_body_cell(ws, row, 1, i, Alignment(horizontal="center", vertical="top"))
        apply_body_cell(ws, row, 2, item.get("published_date", ""))
        apply_body_cell(ws, row, 3, item.get("title", ""))
        apply_body_cell(ws, row, 4, item.get("source", ""))
        apply_body_cell(ws, row, 5, item.get("snippet", ""))

        # 链接列：设为超链接
        url = item.get("url", "")
        cell = apply_body_cell(ws, row, 6, url)
        if url.startswith("http"):
            cell.hyperlink = url
            cell.font = LINK_FONT

        apply_body_cell(ws, row, 7, item.get("language", ""),
                        Alignment(horizontal="center", vertical="top"))

    # 汇总行
    total_row = len(news_items) + 2
    cell = ws.cell(row=total_row, column=1, value="合计")
    cell.font = SUMMARY_FONT
    cell.fill = SUMMARY_FILL
    ws.cell(row=total_row, column=2,
            value="共 {} 条新闻".format(len(news_items)))
    ws.cell(row=total_row, column=2).font = SUMMARY_FONT
    ws.cell(row=total_row, column=2).fill = SUMMARY_FILL
    for col in range(3, 8):
        ws.cell(row=total_row, column=col).fill = SUMMARY_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    auto_column_width(ws)

    print("[Excel] Sheet '新闻汇总' — {} 条记录".format(len(news_items)), file=sys.stderr)


# ─── Sheet 2: 风险评分矩阵 ──────────────────────────────────────────────────

def create_risk_matrix_sheet(wb: Workbook, scored_events: list):
    """Sheet2: 风险评分矩阵 — 含条件格式着色"""
    ws = wb.create_sheet("风险评分矩阵")

    headers = [
        "事件ID", "事件标题", "类型", "涉及国家", "受影响航线",
        "严重度", "概率", "紧急度", "航线权重", "来源数",
        "综合评分", "风险等级", "时间敏感度", "影响类型"
    ]
    apply_header_row(ws, headers)

    for i, event in enumerate(scored_events, 1):
        row = i + 1
        apply_body_cell(ws, row, 1, event.get("event_id", ""))
        apply_body_cell(ws, row, 2, event.get("title", ""))
        apply_body_cell(ws, row, 3, event.get("event_type_name", ""))
        apply_body_cell(ws, row, 4, "、".join(event.get("countries", [])))
        apply_body_cell(ws, row, 5, "、".join(event.get("affected_routes", [])))
        apply_body_cell(ws, row, 6, event.get("severity_score", 0),
                        Alignment(horizontal="center", vertical="top"))
        apply_body_cell(ws, row, 7, event.get("probability_score", 0),
                        Alignment(horizontal="center", vertical="top"))
        apply_body_cell(ws, row, 8, event.get("urgency_score", 0),
                        Alignment(horizontal="center", vertical="top"))
        apply_body_cell(ws, row, 9, event.get("route_importance_score", 0),
                        Alignment(horizontal="center", vertical="top"))
        apply_body_cell(ws, row, 10, event.get("source_count", 0),
                        Alignment(horizontal="center", vertical="top"))

        # 综合评分 + 风险等级：条件着色
        score_cell = apply_body_cell(ws, row, 11, event.get("composite_score", 0),
                                     Alignment(horizontal="center", vertical="top"))
        level = event.get("risk_level", "低")
        level_cell = apply_body_cell(ws, row, 12, level,
                                     Alignment(horizontal="center", vertical="top"))
        apply_risk_fill(score_cell, level)
        apply_risk_fill(level_cell, level)

        apply_body_cell(ws, row, 13, event.get("time_sensitivity", ""))
        apply_body_cell(ws, row, 14, event.get("supply_chain_impact", ""))

    # 汇总行：使用 COUNTIF 公式
    total_row = len(scored_events) + 2
    ws.cell(row=total_row, column=1, value="汇总统计")
    ws.cell(row=total_row, column=1).font = SUMMARY_FONT
    ws.cell(row=total_row, column=1).fill = SUMMARY_FILL

    col_l = get_column_letter(12)  # 风险等级列
    range_str = "{}2:{}{}".format(col_l, col_l, len(scored_events) + 1)

    labels = [("高危", 2), ("中等", 3), ("低", 4)]
    for label, offset in labels:
        r = total_row + offset - 2
        ws.cell(row=r, column=11, value="{}数量".format(label))
        ws.cell(row=r, column=11).font = SUMMARY_FONT
        ws.cell(row=r, column=11).fill = SUMMARY_FILL
        ws.cell(row=r, column=12, value='=COUNTIF({},"{}")'.format(range_str, label))
        ws.cell(row=r, column=12).fill = SUMMARY_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    auto_column_width(ws)

    print("[Excel] Sheet '风险评分矩阵' — {} 个事件".format(len(scored_events)), file=sys.stderr)


# ─── Sheet 3: 供应链影响评估 ────────────────────────────────────────────────

def create_impact_sheet(wb: Workbook, impact_analysis: list):
    """Sheet3: 供应链影响评估 — 含影响路径和替代方案"""
    ws = wb.create_sheet("供应链影响评估")

    headers = [
        "受影响节点", "节点类型", "影响程度", "影响路径",
        "当前状态", "替代方案", "额外成本", "预计恢复时间"
    ]
    apply_header_row(ws, headers)

    for i, item in enumerate(impact_analysis, 1):
        row = i + 1
        apply_body_cell(ws, row, 1, item.get("node_name", ""))
        apply_body_cell(ws, row, 2, item.get("node_type", ""))

        impact_level = item.get("impact_level", "")
        level_cell = apply_body_cell(ws, row, 3, impact_level,
                                     Alignment(horizontal="center", vertical="top"))
        # 影响程度着色
        if impact_level in ("严重", "重大"):
            apply_risk_fill(level_cell, "高危")
        elif impact_level in ("中等", "一般"):
            apply_risk_fill(level_cell, "中等")
        elif impact_level in ("轻微", "微弱"):
            apply_risk_fill(level_cell, "低")

        apply_body_cell(ws, row, 4, item.get("impact_path", ""))
        apply_body_cell(ws, row, 5, item.get("current_status", ""))
        apply_body_cell(ws, row, 6, item.get("alternative", ""))
        apply_body_cell(ws, row, 7, item.get("extra_cost", ""))
        apply_body_cell(ws, row, 8, item.get("recovery_time", ""))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    auto_column_width(ws)

    print("[Excel] Sheet '供应链影响评估' — {} 个节点".format(len(impact_analysis)), file=sys.stderr)


# ─── Sheet 4: 应对建议 ──────────────────────────────────────────────────────

def create_recommendations_sheet(wb: Workbook, recommendations: list):
    """Sheet4: 应对建议 — 含优先级和时间框架"""
    ws = wb.create_sheet("应对建议")

    headers = [
        "序号", "风险等级", "建议措施", "详细说明",
        "优先级", "负责部门", "时间框架", "预估成本"
    ]
    apply_header_row(ws, headers)

    for i, rec in enumerate(recommendations, 1):
        row = i + 1
        apply_body_cell(ws, row, 1, i, Alignment(horizontal="center", vertical="top"))

        level = rec.get("risk_level", "")
        level_cell = apply_body_cell(ws, row, 2, level,
                                     Alignment(horizontal="center", vertical="top"))
        apply_risk_fill(level_cell, level)

        apply_body_cell(ws, row, 3, rec.get("action", ""))
        apply_body_cell(ws, row, 4, rec.get("detail", ""))

        priority = rec.get("priority", "")
        priority_cell = apply_body_cell(ws, row, 5, priority,
                                        Alignment(horizontal="center", vertical="top"))
        if priority == "紧急":
            priority_cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="FF4D4F")
        elif priority == "高":
            priority_cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="FA8C16")

        apply_body_cell(ws, row, 6, rec.get("department", ""))
        apply_body_cell(ws, row, 7, rec.get("timeframe", ""))
        apply_body_cell(ws, row, 8, rec.get("estimated_cost", ""))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    auto_column_width(ws)

    print("[Excel] Sheet '应对建议' — {} 条建议".format(len(recommendations)), file=sys.stderr)


# ─── 主流程 ──────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="中东供应链风险 Excel 报表生成器"
    )
    parser.add_argument(
        "--input", required=True,
        help="输入 JSON 文件路径（包含 news_items, scored_events, impact_analysis, recommendations）"
    )
    parser.add_argument(
        "--output", required=True,
        help="输出 .xlsx 文件路径"
    )
    args = parser.parse_args()

    # 读取输入数据
    try:
        with open(args.input, "r", encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        print("错误: 输入文件不存在 — {}".format(args.input), file=sys.stderr)
        sys.exit(1)
    except json.JSONDecodeError as e:
        print("错误: JSON 解析失败 — {}".format(e), file=sys.stderr)
        sys.exit(1)

    news_items = data.get("news_items", [])
    scored_events = data.get("scored_events", data.get("events", []))
    impact_analysis = data.get("impact_analysis", [])
    recommendations = data.get("recommendations", [])

    print("[Excel] 开始生成报表...", file=sys.stderr)
    print("[Excel] 新闻: {} 条, 事件: {} 个, 影响节点: {} 个, 建议: {} 条".format(
        len(news_items), len(scored_events),
        len(impact_analysis), len(recommendations)
    ), file=sys.stderr)

    # 创建工作簿
    wb = Workbook()

    create_news_sheet(wb, news_items)
    create_risk_matrix_sheet(wb, scored_events)
    create_impact_sheet(wb, impact_analysis)
    create_recommendations_sheet(wb, recommendations)

    # 确保输出目录存在
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # 保存
    wb.save(str(output_path))
    print("[完成] 报表已保存至 {}".format(output_path), file=sys.stderr)

    # 输出文件路径到 stdout，供上层脚本捕获
    print(json.dumps({"output_file": str(output_path)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
